from __future__ import annotations

import io
import re
import zipfile
from datetime import datetime
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


NAVY = "0F1C75"
BLUE = "1C73F5"
GREEN = "00AB0D"
AMBER = "FFB500"
RED = "C62828"
LIGHT_BLUE = "EAF2FF"
LIGHT_RED = "FDECEC"
LIGHT_AMBER = "FFF6D9"
LIGHT_GRAY = "F3F5F8"
WHITE = "FFFFFF"
DARK = "172033"

THIN_GRAY = Side(style="thin", color="D7DCE5")
BORDER = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)


def _is_missing(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, (list, dict, tuple, set)):
        return False
    try:
        return bool(pd.isna(value))
    except (TypeError, ValueError):
        return False


def _safe_filename(value: Any, max_length: int = 80) -> str:
    text = str(value or "sin_dato").strip()
    text = re.sub(r"[^A-Za-zÁÉÍÓÚáéíóúÑñ0-9._-]+", "_", text)
    return text.strip("_")[:max_length] or "sin_dato"


def _list_text(value: Any) -> str:
    if isinstance(value, list):
        return ", ".join(str(item) for item in value)
    if _is_missing(value):
        return ""
    return str(value or "")


def build_referral_reason(row: dict[str, Any]) -> str:
    week = row.get("week_number") or ""
    expected = row.get("expected_activities") or 0
    completed = row.get("completed_activities") or 0
    percent = row.get("completion_percentage")
    average = row.get("average_grade")
    inactivity = row.get("inactivity_hours")
    pending = row.get("pending_assignments") or []
    parts = [
        f"Al cierre de la semana {week}, el estudiante debía registrar un mínimo acumulado de "
        f"{expected} actividades y registra {completed} completadas"
        + (f" ({float(percent):.2f} % del avance esperado)." if percent is not None else ".")
    ]
    if pending:
        names = _list_text(pending)
        parts.append(f"Actividades esperadas pendientes: {names}.")
    if not _is_missing(average):
        parts.append(f"Promedio actual en calificaciones: {float(average):.2f} %.")
    if not _is_missing(inactivity):
        hours = float(inactivity)
        if hours >= 24:
            parts.append(f"Última actividad registrada hace {hours / 24:.1f} días ({hours:.0f} horas).")
        else:
            parts.append(f"Última actividad registrada hace {hours:.0f} horas.")
    late = int(row.get("late_count") or 0)
    if late:
        parts.append(f"Registra {late} entrega(s) tardía(s).")
    reasons = row.get("reasons") or []
    if isinstance(reasons, list):
        additional = [str(reason) for reason in reasons if str(reason).strip()]
        if additional:
            parts.append("Hallazgos del motor de riesgo: " + " ".join(additional))
    return " ".join(parts)


def _set_widths(ws, widths: dict[str, float]) -> None:
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def _write_title(ws, title: str, subtitle: str, priority: str) -> None:
    ws.merge_cells("A1:F1")
    ws["A1"] = title
    ws["A1"].font = Font(size=16, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    ws["A2"] = subtitle
    ws["A2"].font = Font(size=10, color=DARK)
    ws["A2"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    ws["A2"].alignment = Alignment(horizontal="center")

    if priority in {"Urgente", "Prioritaria"}:
        ws.merge_cells("A3:F3")
        ws["A3"] = "PRIORIDAD ALTA — ATENCIÓN REQUERIDA"
        ws["A3"].font = Font(size=12, bold=True, color=WHITE)
        ws["A3"].fill = PatternFill("solid", fgColor=RED)
        ws["A3"].alignment = Alignment(horizontal="center")


def _label_value(ws, row_number: int, label: str, value: Any, start_col: int = 1, end_col: int = 6) -> None:
    ws.cell(row=row_number, column=start_col, value=label)
    ws.cell(row=row_number, column=start_col).font = Font(bold=True, color=NAVY)
    ws.cell(row=row_number, column=start_col).fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    ws.cell(row=row_number, column=start_col).border = BORDER
    ws.merge_cells(start_row=row_number, start_column=start_col + 1, end_row=row_number, end_column=end_col)
    cell = ws.cell(row=row_number, column=start_col + 1, value=_list_text(value))
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    cell.border = BORDER
    for column in range(start_col + 2, end_col + 1):
        ws.cell(row=row_number, column=column).border = BORDER


def create_individual_referral(row: dict[str, Any], academic_advisor: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Derivación"
    priority = str(row.get("intervention_priority") or "")
    _write_title(
        ws,
        "FORMATO DE DERIVACIÓN ACADÉMICA A BIENESTAR",
        f"AVE — Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        priority,
    )
    start = 5 if priority in {"Urgente", "Prioritaria"} else 4
    fields = [
        ("Nombre completo", row.get("student_name")),
        ("Carné", row.get("carne")),
        ("Correo", row.get("email")),
        ("Carrera", row.get("career") or "Pendiente de completar"),
        ("Curso", row.get("course_name")),
        ("Sección", row.get("section_name")),
        ("Semana analizada", f"{row.get('week_number')} de {row.get('total_weeks', 5)}"),
        ("Nivel de riesgo", row.get("overall_risk")),
        ("Prioridad", priority),
        ("Asesor académico", academic_advisor),
        ("Asesor de bienestar", row.get("asesor_bienestar") or row.get("advisor_name") or "Sin asignar"),
        ("Promedio actual", f"{float(row['average_grade']):.2f} %" if not _is_missing(row.get("average_grade")) else "Sin datos"),
        ("Avance esperado", f"{row.get('expected_activities')} actividades"),
        ("Avance registrado", f"{row.get('completed_activities')} actividades"),
        ("Entregas tardías", row.get("late_count") or 0),
        ("Última actividad", row.get("last_activity_at") or "Sin datos"),
    ]
    for offset, (label, value) in enumerate(fields):
        _label_value(ws, start + offset, label, value)

    reason_row = start + len(fields) + 1
    ws.merge_cells(start_row=reason_row, start_column=1, end_row=reason_row, end_column=6)
    ws.cell(reason_row, 1, "RAZÓN DETALLADA DE LA DERIVACIÓN")
    ws.cell(reason_row, 1).font = Font(bold=True, color=WHITE)
    ws.cell(reason_row, 1).fill = PatternFill("solid", fgColor=BLUE)
    ws.cell(reason_row, 1).alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=reason_row + 1, start_column=1, end_row=reason_row + 4, end_column=6)
    reason_cell = ws.cell(reason_row + 1, 1, build_referral_reason(row))
    reason_cell.alignment = Alignment(wrap_text=True, vertical="top")
    reason_cell.border = BORDER

    action_row = reason_row + 6
    _label_value(ws, action_row, "Acciones realizadas", row.get("previous_actions") or "Mensaje de seguimiento académico y revisión de indicadores.")
    _label_value(ws, action_row + 1, "Fecha sugerida de seguimiento", row.get("followup_date") or "Dentro de las próximas 48 horas")
    _label_value(ws, action_row + 2, "Observaciones", row.get("referral_notes") or "")

    _set_widths(ws, {"A": 24, "B": 18, "C": 18, "D": 18, "E": 18, "F": 18})
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.oddFooter.center.text = "AVE — Seguimiento académico y bienestar"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def create_consolidated_referrals(group: pd.DataFrame, advisor_name: str, academic_advisor: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Listado general"
    ws.merge_cells("A1:N1")
    ws["A1"] = f"DERIVACIONES PARA {advisor_name.upper()}"
    ws["A1"].font = Font(size=16, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:N2")
    ws["A2"] = f"Asesor académico: {academic_advisor} | Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    ws["A2"].alignment = Alignment(horizontal="center")

    headers = [
        "Carné",
        "Nombre completo",
        "Correo",
        "Carrera",
        "Curso",
        "Sección",
        "Semana",
        "Riesgo",
        "Prioridad",
        "Esperadas",
        "Completadas",
        "Promedio",
        "Inactividad (h)",
        "Razón resumida",
    ]
    header_row = 4
    for column, header in enumerate(headers, start=1):
        cell = ws.cell(header_row, column, header)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    for row_number, row in enumerate(group.to_dict(orient="records"), start=header_row + 1):
        values = [
            row.get("carne"),
            row.get("student_name"),
            row.get("email"),
            row.get("career"),
            row.get("course_name"),
            row.get("section_name"),
            row.get("week_number"),
            row.get("overall_risk"),
            row.get("intervention_priority"),
            row.get("expected_activities"),
            row.get("completed_activities"),
            row.get("average_grade"),
            row.get("inactivity_hours"),
            build_referral_reason(row)[:500],
        ]
        for column, value in enumerate(values, start=1):
            cell = ws.cell(row_number, column, value)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=column in {2, 3, 4, 5, 6, 14})
            if row.get("overall_risk") == "Alto":
                cell.fill = PatternFill("solid", fgColor=LIGHT_RED)
            elif row.get("overall_risk") == "Moderado":
                cell.fill = PatternFill("solid", fgColor=LIGHT_AMBER)

    end_row = header_row + len(group)
    if end_row >= header_row + 1:
        table = Table(displayName="TablaDerivaciones", ref=f"A{header_row}:N{end_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    widths = [12, 32, 30, 24, 30, 16, 10, 12, 14, 12, 13, 12, 16, 70]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A{header_row}:N{max(end_row, header_row)}"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generate_referral_package(
    selected: pd.DataFrame,
    *,
    academic_advisor: str,
) -> tuple[bytes, list[dict[str, Any]]]:
    if selected.empty:
        raise ValueError("No se seleccionaron estudiantes para derivación.")

    buffer = io.BytesIO()
    records: list[dict[str, Any]] = []
    with zipfile.ZipFile(buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as package:
        for advisor_name, group in selected.groupby("asesor_bienestar", dropna=False):
            advisor = str(advisor_name or "Sin asignar")
            folder = _safe_filename(advisor)
            consolidated = create_consolidated_referrals(group, advisor, academic_advisor)
            package.writestr(f"{folder}/Informe_general.xlsx", consolidated)

            for row in group.to_dict(orient="records"):
                individual = create_individual_referral(row, academic_advisor)
                filename = f"Derivacion_{_safe_filename(row.get('carne'))}_{_safe_filename(row.get('student_name'))}.xlsx"
                package.writestr(f"{folder}/{filename}", individual)
                records.append(
                    {
                        "carne": str(row.get("carne") or ""),
                        "canvas_user_id": str(row.get("canvas_user_id") or ""),
                        "student_name": row.get("student_name"),
                        "email": row.get("email"),
                        "course_id": str(row.get("course_id") or ""),
                        "course_name": row.get("course_name"),
                        "section_name": row.get("section_name"),
                        "week_number": row.get("week_number"),
                        "average_grade": row.get("average_grade"),
                        "completion_percentage": row.get("completion_percentage"),
                        "advisor_name": advisor,
                        "risk_level": row.get("overall_risk"),
                        "priority": row.get("intervention_priority"),
                        "reason": build_referral_reason(row),
                        "status": "generated",
                    }
                )
    return buffer.getvalue(), records
